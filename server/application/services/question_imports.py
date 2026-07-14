from __future__ import annotations

import csv
from collections import Counter
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


QUESTION_IMPORT_HEADERS = ["章节", "题型", "题干", "选项A", "选项B", "选项C", "选项D", "正确答案", "解析", "知识点", "难度", "分值"]
QUESTION_TYPES = {"单项选择题", "多项选择题", "判断题", "填空题", "简答题", "计算题"}
DIFFICULTIES = {"基础", "中等", "困难"}
FORMULA_PREFIXES = ("=", "+", "-", "@")


def build_question_import_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "题库导入"
    sheet.append(QUESTION_IMPORT_HEADERS)
    sheet.append(["第3章 桩基础", "单项选择题", "单桩竖向承载力由哪些部分组成？", "桩侧阻力和桩端阻力", "仅桩端阻力", "仅桩侧阻力", "基础自重", "A", "考查荷载传递", "单桩承载力", "基础", 10])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="075BCF")
    widths = [18, 16, 42, 28, 24, 24, 24, 14, 36, 22, 12, 10]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _string(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _xlsx_rows(content: bytes) -> list[tuple[int, dict[str, Any], set[str]]]:
    workbook = load_workbook(BytesIO(content), data_only=False, read_only=True)
    sheet = workbook.active
    cells = list(sheet.iter_rows())
    if not cells:
        return []
    headers = [_string(cell.value) for cell in cells[0]]
    if headers[:len(QUESTION_IMPORT_HEADERS)] != QUESTION_IMPORT_HEADERS:
        raise ValueError("题库模板表头不正确，请下载最新模板")
    result = []
    for row_number, row in enumerate(cells[1:], start=2):
        values = {header: row[index].value if index < len(row) else None for index, header in enumerate(QUESTION_IMPORT_HEADERS)}
        if not any(_string(value) for value in values.values()):
            continue
        formulas = {header for index, header in enumerate(QUESTION_IMPORT_HEADERS) if index < len(row) and row[index].data_type == "f"}
        result.append((row_number, values, formulas))
    return result


def _csv_rows(content: bytes) -> list[tuple[int, dict[str, Any], set[str]]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("CSV 文件必须使用 UTF-8 编码") from exc
    reader = csv.DictReader(StringIO(text))
    if not reader.fieldnames or reader.fieldnames[:len(QUESTION_IMPORT_HEADERS)] != QUESTION_IMPORT_HEADERS:
        raise ValueError("题库模板表头不正确，请下载最新模板")
    result = []
    for row_number, source in enumerate(reader, start=2):
        values = {header: source.get(header) for header in QUESTION_IMPORT_HEADERS}
        if not any(_string(value) for value in values.values()):
            continue
        formulas = {header for header, value in values.items() if _string(value).startswith(FORMULA_PREFIXES)}
        result.append((row_number, values, formulas))
    return result


def _error(row: int, field: str, code: str, reason: str) -> dict:
    return {"row": row, "field": field, "code": code, "reason": reason}


def parse_question_import(filename: str, content: bytes) -> dict:
    suffix = Path(filename).suffix.lower()
    if suffix == ".xlsx":
        raw_rows = _xlsx_rows(content)
    elif suffix == ".csv":
        raw_rows = _csv_rows(content)
    else:
        raise ValueError("仅支持 XLSX 或 UTF-8 CSV 文件")

    stems = [_string(values["题干"]) for _row, values, _formulas in raw_rows]
    duplicate_stems = {stem for stem, count in Counter(stems).items() if stem and count > 1}
    normalized_rows: list[dict] = []
    errors: list[dict] = []

    for row_number, values, formulas in raw_rows:
        row_errors: list[dict] = []
        for field in formulas:
            row_errors.append(_error(row_number, field, "FORMULA_NOT_ALLOWED", f"{field}不允许包含公式"))
        text = _string(values["题干"])
        question_type = _string(values["题型"])
        chapter = _string(values["章节"])
        answer = _string(values["正确答案"])
        difficulty = _string(values["难度"]) or "基础"
        if not text:
            row_errors.append(_error(row_number, "题干", "REQUIRED_FIELD", "题干不能为空"))
        if text in duplicate_stems:
            row_errors.append(_error(row_number, "题干", "DUPLICATE_IN_FILE", "文件内题干重复"))
        if question_type not in QUESTION_TYPES:
            row_errors.append(_error(row_number, "题型", "INVALID_TYPE", "题型不在允许范围内"))
        if not chapter:
            row_errors.append(_error(row_number, "章节", "REQUIRED_FIELD", "章节不能为空"))
        if not answer:
            row_errors.append(_error(row_number, "正确答案", "ANSWER_REQUIRED", "正确答案不能为空"))
        if difficulty not in DIFFICULTIES:
            row_errors.append(_error(row_number, "难度", "INVALID_DIFFICULTY", "难度必须为基础、中等或困难"))
        try:
            points = float(values["分值"] or 10)
            if points <= 0 or points > 1000:
                raise ValueError
        except (TypeError, ValueError):
            points = 10
            row_errors.append(_error(row_number, "分值", "INVALID_POINTS", "分值必须大于 0 且不超过 1000"))

        options = [
            {"label": label, "text": _string(values[f"选项{label}"])}
            for label in "ABCD" if _string(values[f"选项{label}"])
        ]
        if question_type in {"单项选择题", "多项选择题"}:
            if len(options) < 2:
                row_errors.append(_error(row_number, "选项", "OPTIONS_REQUIRED", "选择题至少需要两个选项"))
            answer_labels = [part.strip().upper() for part in answer.replace("，", ",").split(",") if part.strip()]
            valid_labels = {item["label"] for item in options}
            if answer and (not answer_labels or any(label not in valid_labels for label in answer_labels)):
                row_errors.append(_error(row_number, "正确答案", "INVALID_ANSWER", "选择题答案必须对应有效选项"))
            correct_answer: Any = answer_labels if question_type == "多项选择题" else (answer_labels[0] if answer_labels else answer)
        else:
            options = []
            correct_answer = answer

        if row_errors:
            errors.extend(row_errors)
            continue
        normalized_rows.append({
            "text": text, "questionType": question_type, "options": options,
            "correctAnswer": correct_answer, "explanation": _string(values["解析"]) or None,
            "rubric": [], "difficulty": difficulty, "points": points, "chapter": chapter,
            "knowledgePoint": _string(values["知识点"]) or None,
        })

    return {
        "rows": normalized_rows,
        "errors": errors,
        "summary": {"total": len(raw_rows), "valid": len(normalized_rows), "errors": len(errors)},
    }
