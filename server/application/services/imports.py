from __future__ import annotations

import csv
import io
from typing import Any

from openpyxl import Workbook, load_workbook


DANGEROUS_PREFIXES = ("=", "+", "-", "@")
REQUIRED_COLUMNS = ("姓名", "学号", "班级")


def dangerous(value: Any) -> bool:
    return isinstance(value, str) and value.lstrip().startswith(DANGEROUS_PREFIXES)


def validate_student_rows(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    valid: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    seen: set[str] = set()
    for index, row in enumerate(rows, start=2):
        name = str(row.get("姓名") or row.get("name") or "").strip()
        student_no = str(row.get("学号") or row.get("studentNo") or "").strip()
        class_name = str(row.get("班级") or row.get("className") or "").strip()
        values = (name, student_no, class_name)
        if any(dangerous(value) for value in values):
            errors.append({"row": index, "studentNo": student_no, "code": "FORMULA_NOT_ALLOWED", "reason": "不允许包含公式内容"})
            continue
        missing = next((label for label, value in zip(REQUIRED_COLUMNS, values) if not value), None)
        if missing:
            errors.append({"row": index, "studentNo": student_no, "code": "REQUIRED_FIELD", "reason": f"{missing}不能为空"})
            continue
        if student_no in seen:
            errors.append({"row": index, "studentNo": student_no, "code": "DUPLICATE_IN_FILE", "reason": "文件内学号重复"})
            continue
        seen.add(student_no)
        valid.append({"name": name, "studentNo": student_no, "username": student_no, "className": class_name})
    return {"valid": valid, "errors": errors}


def parse_pasted_students(text: str) -> dict[str, list[dict[str, Any]]]:
    lines = [line for line in text.splitlines() if line.strip()]
    if not lines:
        return {"valid": [], "errors": []}
    dialect = csv.excel_tab if "\t" in lines[0] else csv.excel
    return validate_student_rows(list(csv.DictReader(lines, dialect=dialect)))


def parse_student_file(content: bytes, filename: str) -> dict[str, list[dict[str, Any]]]:
    if filename.lower().endswith(".csv"):
        return validate_student_rows(list(csv.DictReader(io.StringIO(content.decode("utf-8-sig")))))
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
    sheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows())]
    rows = []
    formula_rows: set[int] = set()
    for row_number, cells in enumerate(sheet.iter_rows(), start=2):
        values = []
        for cell in cells:
            if cell.data_type == "f":
                formula_rows.add(row_number)
            values.append(cell.value)
        rows.append(dict(zip(headers, values)))
    result = validate_student_rows(rows)
    for row_number in sorted(formula_rows):
        result["valid"] = [item for item in result["valid"] if item.get("studentNo") != str(rows[row_number - 2].get("学号") or "")]
        result["errors"].append({"row": row_number, "studentNo": str(rows[row_number - 2].get("学号") or ""), "code": "FORMULA_NOT_ALLOWED", "reason": "不允许包含 Excel 公式"})
    return result


def build_import_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "学生导入"
    sheet.append(["姓名", "学号", "班级", "登录账号", "初始密码"])
    sheet.append(["示例学生", "20260001", "土木工程2401班", "20260001", "Student-123"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
