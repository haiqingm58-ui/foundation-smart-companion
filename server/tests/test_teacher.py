from __future__ import annotations

from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import select


QUESTION_IMPORT_HEADERS = ["章节", "题型", "题干", "选项A", "选项B", "选项C", "选项D", "正确答案", "解析", "知识点", "难度", "分值"]


def question_workbook(rows: list[list]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(QUESTION_IMPORT_HEADERS)
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


@pytest.fixture()
def teacher_context(database_url: str, tmp_path: Path):
    from server.application.config import Settings
    from server.application.database import create_database
    from server.application.main import create_app
    from server.application.migrations import upgrade_database
    from server.application.models import ClassRoom, SessionToken, Student, Teacher, TeacherStudentBinding, User
    from server.application.security import hash_password, token_digest

    upgrade_database(database_url)
    database = create_database(database_url)
    settings = Settings(
        database_url=database_url, secret_key="test-secret", data_dir=tmp_path, upload_dir=tmp_path / "uploads",
        session_ttl_seconds=3600, captcha_ttl_seconds=120, cookie_secure=False, cookie_path="/",
        llm_api_url="", llm_api_key="", llm_model="test",
    )
    settings.upload_dir.mkdir(parents=True, exist_ok=True)
    password = hash_password("Teacher-123")
    with database.session() as session:
        users = [
            User(id="teacher-user-1", username="teacher1", password_hash=password, password_algorithm="argon2", role="teacher", role_label="指导老师", name="教师一", student_no="T1", college="土木工程学院", school="湖南大学", status="active"),
            User(id="teacher-user-2", username="teacher2", password_hash=password, password_algorithm="argon2", role="teacher", role_label="指导老师", name="教师二", student_no="T2", college="土木工程学院", school="湖南大学", status="active"),
            User(id="student-user-1", username="student1", password_hash=password, password_algorithm="argon2", role="student", role_label="学生", name="学生一", student_no="S1", college="土木工程学院", school="湖南大学", status="active"),
            User(id="student-user-2", username="student2", password_hash=password, password_algorithm="argon2", role="student", role_label="学生", name="学生二", student_no="S2", college="土木工程学院", school="湖南大学", status="active"),
        ]
        session.add_all(users)
        session.flush()
        classroom = ClassRoom(id="class-1", name="土木工程2401班", grade="2024", major="土木工程", college="土木工程学院")
        teachers = [
            Teacher(id="teacher-1", user_id="teacher-user-1", teacher_no="T1", college="土木工程学院"),
            Teacher(id="teacher-2", user_id="teacher-user-2", teacher_no="T2", college="土木工程学院"),
        ]
        students = [
            Student(id="student-1", user_id="student-user-1", student_no="S1", class_id="class-1", progress=45, average_score=82),
            Student(id="student-2", user_id="student-user-2", student_no="S2", class_id="class-1", progress=20, average_score=66),
        ]
        session.add(classroom)
        session.add_all(teachers)
        session.add_all(students)
        session.flush()
        session.add_all(
            [
                TeacherStudentBinding(id="binding-1", teacher_id="teacher-1", student_id="student-1", class_id="class-1", status="active", created_by="teacher-user-1"),
                TeacherStudentBinding(id="binding-2", teacher_id="teacher-2", student_id="student-2", class_id="class-1", status="active", created_by="teacher-user-2"),
                SessionToken(id="teacher-session", user_id="teacher-user-1", token_hash=token_digest("teacher-token"), csrf_hash=token_digest("teacher-csrf"), expires_at=datetime.now(timezone.utc) + timedelta(hours=1)),
            ]
        )
        session.commit()
    app = create_app(settings=settings, database=database)
    client = TestClient(app)
    client.cookies.set("foundation_session", "teacher-token")
    client.cookies.set("foundation_csrf", "teacher-csrf")
    return client, database, settings


def test_teacher_only_sees_bound_students(teacher_context) -> None:
    client, _database, _settings = teacher_context
    response = client.get("/api/teacher/students?page=1&pageSize=20")
    assert response.status_code == 200
    items = response.json()["data"]["items"]
    assert [item["studentNo"] for item in items] == ["S1"]
    assert client.get("/api/teacher/students/student-2").status_code == 404


def test_teacher_uploads_real_markdown_resource(teacher_context) -> None:
    from server.application.models import KnowledgeChunk, Resource

    client, database, settings = teacher_context
    response = client.post(
        "/api/teacher/resources",
        data={"chapter": "第3章 桩基础", "knowledgePoint": "桩侧阻力", "visibility": "class"},
        files={"file": ("pile-notes.md", b"# Pile\nThe pile side resistance transfers load.", "text/markdown")},
        headers={"X-CSRF-Token": "teacher-csrf"},
    )
    assert response.status_code == 200
    resource_id = response.json()["data"]["id"]
    assert client.get(f"/api/teacher/resources/{resource_id}/preview").status_code == 200
    with database.session() as session:
        resource = session.get(Resource, resource_id)
        assert Path(resource.storage_path).is_file()
        assert Path(resource.storage_path).resolve().is_relative_to(settings.upload_dir.resolve())
        assert session.scalar(select(KnowledgeChunk).where(KnowledgeChunk.resource_id == resource_id)) is not None


def test_teacher_manages_question_and_assignment_for_own_student(teacher_context) -> None:
    client, _database, _settings = teacher_context
    question = client.post(
        "/api/teacher/questions",
        json={"text": "桩侧阻力主要受哪些因素影响？", "questionType": "简答题", "difficulty": "基础", "points": 10, "chapter": "第3章 桩基础", "knowledgePoint": "桩侧阻力", "options": [], "correctAnswer": "土性与相对位移", "explanation": "考查荷载传递"},
        headers={"X-CSRF-Token": "teacher-csrf"},
    )
    assert question.status_code == 200
    question_data = question.json()["data"]
    assert question_data["subjectId"] == "foundation-engineering"
    assert len(question_data["knowledgePoints"]) == 1
    assert question_data["status"] == "active"
    question_id = question_data["id"]
    assignment = client.post(
        "/api/teacher/assignments",
        json={"title": "桩基础练习", "description": "完成简答题", "studentIds": ["student-1"], "questionIds": [question_id], "totalPoints": 10, "allowResubmit": False, "autoGrade": False, "status": "published"},
        headers={"X-CSRF-Token": "teacher-csrf"},
    )
    assert assignment.status_code == 200
    assignment_list = client.get("/api/teacher/assignments")
    assert assignment_list.json()["data"]["items"][0]["targetCount"] == 1
    assert assignment_list.json()["data"]["items"][0]["completionRate"] == 0
    forbidden_target = client.post(
        "/api/teacher/assignments",
        json={"title": "越权作业", "studentIds": ["student-2"], "questionIds": [question_id], "totalPoints": 10, "status": "published"},
        headers={"X-CSRF-Token": "teacher-csrf"},
    )
    assert forbidden_target.status_code == 403


def test_teacher_dashboard_uses_real_owned_data(teacher_context) -> None:
    client, _database, _settings = teacher_context
    response = client.get("/api/teacher/dashboard")
    assert response.status_code == 200
    data = response.json()["data"]
    assert data["studentTotal"] == 1
    assert data["averageScore"] == 82


def test_teacher_publishes_and_lists_notices(teacher_context) -> None:
    client, _database, _settings = teacher_context
    created = client.post(
        "/api/teacher/notices",
        json={"title": "本周答疑安排", "content": "周三下午在岩土楼答疑。", "audience": "class", "classScope": ["class-1"]},
        headers={"X-CSRF-Token": "teacher-csrf"},
    )
    assert created.status_code == 200
    response = client.get("/api/teacher/notices")
    assert response.status_code == 200
    assert response.json()["data"]["items"][0]["title"] == "本周答疑安排"


def test_teacher_grades_only_owned_submission(teacher_context) -> None:
    from server.application.models import Assignment, AssignmentTarget, Submission

    client, database, _settings = teacher_context
    now = datetime.now(timezone.utc)
    with database.session() as session:
        session.add(Assignment(id="assignment-owned", title="作业一", teacher_id="teacher-1", total_points=100, status="published"))
        session.add(Assignment(id="assignment-other", title="作业二", teacher_id="teacher-2", total_points=100, status="published"))
        session.add(AssignmentTarget(id="target-owned", assignment_id="assignment-owned", student_id="student-1", class_id="class-1"))
        session.add(AssignmentTarget(id="target-other", assignment_id="assignment-other", student_id="student-2", class_id="class-1"))
        session.add(Submission(id="submission-owned", assignment_id="assignment-owned", student_id="student-1", submitted_at=now, status="submitted"))
        session.add(Submission(id="submission-other", assignment_id="assignment-other", student_id="student-2", submitted_at=now, status="submitted"))
        session.commit()

    response = client.put(
        "/api/teacher/submissions/submission-owned/grade",
        json={"score": 88, "feedback": "步骤完整，注意单位。"},
        headers={"X-CSRF-Token": "teacher-csrf"},
    )
    assert response.status_code == 200
    assert client.put(
        "/api/teacher/submissions/submission-other/grade",
        json={"score": 90, "feedback": "越权"},
        headers={"X-CSRF-Token": "teacher-csrf"},
    ).status_code == 404
    listed = client.get("/api/teacher/submissions")
    assert listed.status_code == 200
    assert listed.json()["data"]["items"][0]["score"] == 88


def test_teacher_reads_rich_owned_submission_snapshot_for_manual_grading(teacher_context) -> None:
    from server.application.models import Assignment, AssignmentQuestion, AssignmentTarget, Question, Submission, SubmissionAnswer

    client, database, _settings = teacher_context
    now = datetime.now(timezone.utc)
    snapshot = {
        "id": "grading-question", "subjectId": "soil-mechanics", "text": "说明渗透系数的影响因素。",
        "questionType": "简答题", "options": [], "correctAnswer": None, "explanation": "考虑土粒级配和孔隙比。",
        "rubric": [{"criterion": "指出级配", "points": 8}, {"criterion": "指出孔隙比", "points": 8}],
        "difficulty": "中等", "chapter": "第二章 土的渗透性", "knowledgePointIds": ["soil-permeability"],
        "attachments": [], "gradingMode": "manual", "sectionTitle": "二、简答题", "sequence": 1, "points": 20,
    }
    with database.session() as session:
        session.add(Question(id="grading-question", text=snapshot["text"], question_type="简答题", options=[], correct_answer=None, rubric=snapshot["rubric"], difficulty="中等", points=20, subject_id="soil-mechanics", grading_mode="manual", status="active", source="teacher"))
        session.add(Assignment(id="grading-assignment", title="土力学期中测验", teacher_id="teacher-1", total_points=20, status="published"))
        session.flush()
        session.add(AssignmentQuestion(id="grading-aq", assignment_id="grading-assignment", question_id="grading-question", sequence=1, points=20, question_snapshot=snapshot))
        session.add(AssignmentTarget(id="grading-target", assignment_id="grading-assignment", student_id="student-1", class_id="class-1"))
        session.flush()
        session.add(Submission(id="grading-submission", assignment_id="grading-assignment", student_id="student-1", submitted_at=now, status="pending_review"))
        session.flush()
        session.add(SubmissionAnswer(id="grading-answer", submission_id="grading-submission", question_id="grading-question", answer="与土粒级配和孔隙比有关。", score=None, criteria_scores={}, confidence=0.52))
        session.commit()

    response = client.get("/api/teacher/submissions/grading-submission")
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["totalPoints"] == 20
    assert data["studentName"] == "学生一"
    assert data["questions"][0]["text"] == snapshot["text"]
    assert data["questions"][0]["answer"] == "与土粒级配和孔隙比有关。"
    assert data["questions"][0]["rubric"] == snapshot["rubric"]
    assert data["questions"][0]["confidence"] == 0.52


def test_teacher_grades_each_submission_answer_and_can_override_final_total(teacher_context) -> None:
    from server.application.models import Assignment, AssignmentQuestion, AssignmentTarget, Question, Submission, SubmissionAnswer

    client, database, _settings = teacher_context
    now = datetime.now(timezone.utc)
    with database.session() as session:
        session.add(Question(id="answer-grade-question", text="说明固结过程。", question_type="简答题", options=[], correct_answer=None, rubric=[{"criterion": "过程", "points": 20}], difficulty="中等", points=20, subject_id="soil-mechanics", grading_mode="manual", status="active", source="teacher"))
        session.add(Assignment(id="answer-grade-assignment", title="固结测验", teacher_id="teacher-1", total_points=20, status="published"))
        session.flush()
        session.add(AssignmentQuestion(id="answer-grade-aq", assignment_id="answer-grade-assignment", question_id="answer-grade-question", sequence=1, points=20, question_snapshot={"id": "answer-grade-question", "text": "说明固结过程。", "questionType": "简答题", "rubric": [{"criterion": "过程", "points": 20}], "gradingMode": "manual", "points": 20, "sequence": 1}))
        session.add(AssignmentTarget(id="answer-grade-target", assignment_id="answer-grade-assignment", student_id="student-1", class_id="class-1"))
        session.flush()
        session.add(Submission(id="answer-grade-submission", assignment_id="answer-grade-assignment", student_id="student-1", submitted_at=now, status="pending_review"))
        session.flush()
        session.add(SubmissionAnswer(id="answer-grade-answer", submission_id="answer-grade-submission", question_id="answer-grade-question", answer="孔隙水排出，有效应力增加。", score=None, criteria_scores={}, confidence=0.45))
        session.commit()

    response = client.put(
        "/api/teacher/submissions/answer-grade-submission/grade",
        json={
            "score": 18, "feedback": "结论正确，过程可更完整。",
            "answers": [{"questionId": "answer-grade-question", "score": 17, "criteriaScores": {"过程": 17}, "feedback": "补充时间效应。"}],
        },
        headers={"X-CSRF-Token": "teacher-csrf"},
    )
    assert response.status_code == 200, response.text
    with database.session() as session:
        submission = session.get(Submission, "answer-grade-submission")
        answer = session.get(SubmissionAnswer, "answer-grade-answer")
        assert (submission.score, submission.feedback, submission.status) == (18, "结论正确，过程可更完整。", "graded")
        assert (answer.score, answer.criteria_scores, answer.feedback) == (17, {"过程": 17}, "补充时间效应。")


def test_teacher_downloads_template_and_previews_question_import(teacher_context) -> None:
    client, _database, _settings = teacher_context
    template = client.get("/api/teacher/question-import-template")
    assert template.status_code == 200
    assert "attachment" in template.headers["content-disposition"]
    workbook = load_workbook(BytesIO(template.content), read_only=True)
    assert [cell.value for cell in next(workbook.active.iter_rows())] == QUESTION_IMPORT_HEADERS

    preview = client.post(
        "/api/teacher/questions/import-preview",
        files={"file": ("questions.xlsx", question_workbook([
            ["第3章 桩基础", "单项选择题", "单桩竖向承载力由哪些部分组成？", "桩侧阻力和桩端阻力", "仅桩端阻力", "仅桩侧阻力", "基础自重", "A", "考查荷载传递", "单桩承载力", "基础", 10],
        ]), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"X-CSRF-Token": "teacher-csrf"},
    )
    assert preview.status_code == 200
    data = preview.json()["data"]
    assert data["summary"] == {"total": 1, "valid": 1, "errors": 0}
    assert data["rows"][0]["questionType"] == "单项选择题"
    assert data["rows"][0]["options"][0] == {"label": "A", "text": "桩侧阻力和桩端阻力"}
    assert data["rows"][0]["correctAnswer"] == "A"


def test_question_import_reports_rows_and_inserts_transactionally(teacher_context) -> None:
    from server.application.models import Question

    client, database, _settings = teacher_context
    invalid = client.post(
        "/api/teacher/questions/import-preview",
        files={"file": ("bad.xlsx", question_workbook([
            ["第3章 桩基础", "未知题型", "重复题干", "A项", "B项", "", "", "A", "", "桩基础", "基础", 10],
            ["第3章 桩基础", "单项选择题", "重复题干", "A项", "B项", "", "", "", "", "桩基础", "基础", 10],
            ["第3章 桩基础", "简答题", "=1+1", "", "", "", "", "答案", "", "桩基础", "基础", 10],
        ]), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"X-CSRF-Token": "teacher-csrf"},
    )
    assert invalid.status_code == 200
    errors = invalid.json()["data"]["errors"]
    assert {item["code"] for item in errors} >= {"INVALID_TYPE", "DUPLICATE_IN_FILE", "ANSWER_REQUIRED", "FORMULA_NOT_ALLOWED"}

    rows = [
        {"text": "桩侧负摩阻力在什么条件下产生？", "questionType": "简答题", "options": [], "correctAnswer": "土体相对桩身向下位移", "explanation": "考查相对位移", "rubric": [], "difficulty": "中等", "points": 12, "chapter": "第3章 桩基础", "knowledgePoint": "负摩阻力"},
        {"text": "地基承载力验算的基本要求是什么？", "questionType": "简答题", "options": [], "correctAnswer": "基底压力不超过承载力特征值", "explanation": "", "rubric": [], "difficulty": "基础", "points": 10, "chapter": "第2章 浅基础", "knowledgePoint": "地基承载力"},
    ]
    imported = client.post(
        "/api/teacher/questions/import",
        json={"rows": rows}, headers={"X-CSRF-Token": "teacher-csrf"},
    )
    assert imported.status_code == 200
    assert imported.json()["data"]["created"] == 2
    with database.session() as session:
        owned = session.scalars(select(Question).where(Question.created_by == "teacher-user-1")).all()
        assert len(owned) == 2
        assert all(item.source == "teacher-import" for item in owned)

    rejected = client.post(
        "/api/teacher/questions/import",
        json={"rows": [rows[0], {**rows[1], "questionType": "未知题型"}]},
        headers={"X-CSRF-Token": "teacher-csrf"},
    )
    assert rejected.status_code == 422
    with database.session() as session:
        assert len(session.scalars(select(Question).where(Question.created_by == "teacher-user-1")).all()) == 2


def test_question_import_rejects_missing_legacy_knowledge_point_without_creating_question(teacher_context) -> None:
    from server.application.models import Question

    client, database, _settings = teacher_context
    response = client.post(
        "/api/teacher/questions/import",
        json={"rows": [{"text": "缺少知识点的旧题目", "questionType": "简答题", "options": [], "correctAnswer": "参考答案", "chapter": "第3章 桩基础"}]},
        headers={"X-CSRF-Token": "teacher-csrf"},
    )
    assert response.status_code == 422
    assert response.json()["code"] == "KNOWLEDGE_POINT_REQUIRED"
    with database.session() as session:
        assert session.scalars(select(Question).where(Question.created_by == "teacher-user-1")).all() == []
